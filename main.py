from fastapi import FastAPI, File, UploadFile, BackgroundTasks
from fastapi.responses import StreamingResponse
import pandas as pd
import shutil
import os
from uuid import uuid4
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, numbers
from io import BytesIO
from urllib.parse import quote

app = FastAPI()

def normalize_code(value):
    try:
        return str(int(float(value))).strip()
    except:
        return str(value).strip()

@app.post("/merge-sc-monthlyp/")
async def merge_sc_monthlyp(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    unique_id = uuid4().hex
    temp_input_path = f"/tmp/{unique_id}_{file.filename}"

    with open(temp_input_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        excel_data = pd.read_excel(temp_input_path, sheet_name=None)
        sheet1 = excel_data.get("Sheet1")
        if sheet1 is None:
            return {"error": "Sheet1 시트를 찾을 수 없습니다."}

        rival_df = pd.read_excel(temp_input_path, sheet_name="Rival", header=26)

        sheet1.columns = [str(c).strip() for c in sheet1.columns]
        code_col = next((col for col in sheet1.columns if col.strip().lower() == 'code'), None)
        monthlyp_col = next((col for col in sheet1.columns if '월초p' in col.strip().lower()), None)
        lump_col = next((col for col in sheet1.columns if '일시납' in col.strip().lower()), None)
        if not code_col or not monthlyp_col or not lump_col:
            return {"error": "Sheet1에 'Code', '월초P', 또는 '일시납' 컬럼이 없습니다."}

        sheet1 = sheet1.dropna(subset=[code_col, monthlyp_col, lump_col]).copy()
        sheet1[code_col] = sheet1[code_col].apply(normalize_code)

        # 쉼표 제거 후 숫자로 변환
        sheet1[lump_col] = sheet1[lump_col].astype(str).str.replace(",", "").astype(float)

        sheet1["lump_bonus"] = sheet1[lump_col] / 200
        sheet1["total_calc"] = sheet1[monthlyp_col] + sheet1["lump_bonus"]
        code_to_p = sheet1.set_index(code_col)["total_calc"].to_dict()
        code_to_bonus = sheet1.set_index(code_col)["lump_bonus"].to_dict()

        wb = load_workbook(temp_input_path)
        if "Rival" not in wb.sheetnames:
            return {"error": "엑셀 파일에 'Rival' 시트가 없습니다."}

        ws = wb["Rival"]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_font = Font(color="FF0000")

        updated_count = 0

        for idx, row in rival_df.iterrows():
            left_code = normalize_code(row["코드"] if "코드" in row else row.iloc[8])
            right_code = normalize_code(row["코드.1"] if "코드.1" in row else row.iloc[17])
            left_value = code_to_p.get(left_code)
            right_value = code_to_p.get(right_code)
            left_bonus = code_to_bonus.get(left_code)
            right_bonus = code_to_bonus.get(right_code)

            if left_value is not None:
                left_cell = ws.cell(row=idx + 28, column=15)
                left_cell.value = left_value
                left_cell.fill = yellow_fill
                left_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            if right_value is not None:
                right_cell = ws.cell(row=idx + 28, column=24)
                right_cell.value = right_value
                right_cell.fill = yellow_fill
                right_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

            # AA열: 27번 열에 일시납/200 표시
            if left_bonus is not None:
                ws.cell(row=idx + 28, column=27).value = left_bonus
            if right_bonus is not None:
                ws.cell(row=idx + 28, column=28).value = right_bonus

            winner_name = None
            debug_note = ""
            if left_value is not None and right_value is not None:
                if left_value > right_value:
                    ws.cell(row=idx + 28, column=15).font = red_font
                    winner_name = str(row.get("FSR") or "").strip()
                    debug_note = f"{left_value} > {right_value}"
                elif right_value > left_value:
                    ws.cell(row=idx + 28, column=24).font = red_font
                    winner_name = str(row.get("FSR.1") or "").strip()
                    debug_note = f"{right_value} > {left_value}"
                else:
                    debug_note = f"{left_value} = {right_value}"

            if winner_name:
                ws.cell(row=idx + 28, column=25).value = winner_name
            if debug_note:
                ws.cell(row=idx + 28, column=26).value = debug_note

            if left_value is not None:
                updated_count += 1
                print(f"[MATCH-LEFT] Code: {left_code} → {left_value}")
            else:
                print(f"[MISS-LEFT] Code: {left_code}")
            if right_value is not None:
                updated_count += 1
                print(f"[MATCH-RIGHT] Code: {right_code} → {right_value}")
            else:
                print(f"[MISS-RIGHT] Code: {right_code}")

        print(f"[RESULT] Total updated cells: {updated_count}")

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        background_tasks.add_task(cleanup_files, [temp_input_path])

        safe_filename = quote(f"merged_{file.filename}")
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"}
        )

    except Exception as e:
        print(f"[ERROR] {str(e)}")
        return {"error": str(e)}


def cleanup_files(paths):
    for path in paths:
        if os.path.exists(path):
            os.remove(path)
